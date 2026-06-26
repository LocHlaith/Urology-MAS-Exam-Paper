#!/usr/bin/env python3
"""Generate source-data Excel workbooks for the current manuscript panels.

The earlier source-data helper in this repository follows an older figure
numbering scheme.  This script keys off the PDFs that currently exist in
plot/panels and writes one workbook per panel using the requested
source_data_fig_*_*.xlsx naming convention.
"""
from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PLOT_ROOT = Path(__file__).resolve().parents[1]
DERIVED = PLOT_ROOT / "derived_data"
PANELS = PLOT_ROOT / "panels"
SOURCE_DIR = PLOT_ROOT / "data" / "source_data"
SOURCE_DIR.mkdir(parents=True, exist_ok=True)


PANEL_SOURCE_MAP: dict[str, list[str]] = {
    "Figure1A": [
        "item_master.csv",
        "mas_question_generation_time.csv",
        "table1_item_blueprint_textual_characteristics.csv",
    ],
    "Figure1B": [
        "machine_safety_screening_source_summary.csv",
        "machine_safety_screening_by_run.csv",
        "machine_safety_screening_crosswalk.csv",
    ],
    "Figure1C": ["exam_form_assignment.csv"],
    "Figure1D": ["exam_form_assignment.csv", "table2_examinee_baseline_randomization_balance.csv"],
    "Figure1E": ["table3_primary_key_secondary_endpoints.csv"],
    "Figure2B": ["fig2B_quality_difference_item_scores.csv", "fig2B_quality_difference_stats.csv"],
    "Figure2C": [
        "fig2C_dimension_scores_item_scores.csv",
        "fig2C_dimension_scores_stats.csv",
        "fig2C_dimension_score_annotations.csv",
    ],
    "Figure2D": [
        "fig2D_quality_by_cognitive_level_item_scores.csv",
        "fig2D_quality_by_cognitive_level_stats.csv",
    ],
    "Figure2E": [
        "fig2E_expert_quality_interaction_model_input.csv",
        "fig2E_expert_quality_interaction_plot_data.csv",
        "fig2E_expert_quality_interaction_contrasts.csv",
        "fig2E_expert_quality_interaction_fixed_effects.csv",
    ],
    "Figure2F": [
        "fig2F_machine_rating_icc_item_scores.csv",
        "fig2F_machine_rating_icc_stats.csv",
        "fig2F_machine_rating_icc_bootstrap.csv",
    ],
    "Figure3B": ["fig3B_student_correct_rate_raw.csv", "fig3B_student_correct_rate_stats.csv"],
    "Figure3C": [
        "fig3C_student_accuracy_horizontal_stats.csv",
        "fig3C_student_accuracy_student_rates.csv",
    ],
    "Figure3D": [
        "fig3D_adjusted_student_probabilities.csv",
        "fig3D_source_cognitive_interaction_contrasts.csv",
    ],
    "Figure3E": [
        "fig3E_ctt_scatter_item_data.csv",
        "fig3E_ctt_scatter_summary.csv",
        "fig3E_ctt_linear_fit.csv",
    ],
    "Figure3F": ["fig3F_reliability_bootstrap_ci.csv", "responses.csv", "item_master.csv"],
    "Figure4B": ["fig4B_expert_source_identification_accuracy.csv"],
    "Figure4C": ["fig4C_expert_source_confusion_counts.csv"],
    "Figure4D": ["fig4D_expert_guessed_mas_model_input.csv", "fig4D_expert_guessed_mas_model_forest.csv"],
    "Figure4E": ["fig4E_student_source_identification_accuracy.csv", "source_detection.csv"],
    "Figure4F": ["fig4F_mas_timing_sensitivity.csv", "mas_question_generation_time.csv"],
    "Figure5A": ["fig5_workflow_time_by_item_type.csv", "fig5_workflow_summary.csv"],
    "Figure5B": ["fig5_workflow_summary.csv", "fig5_workflow_time_by_item_type.csv"],
    "Figure5C": ["fig5_time_sensitivity.csv", "fig5_workflow_summary.csv"],
    "Figure5D": ["fig5_api_cost_by_stage.csv", "fig5_token_cost_estimate_by_type_stage.csv"],
    "Figure5E": [
        "fig5E_total_cost_comparison.csv",
        "fig5_api_cost_by_stage.csv",
        "ai_efficiency_filled_from_update.csv",
        "fig5_workflow_summary.csv",
    ],
    "Figure6A": ["fig6_fatigue_order_student_data.csv"],
    "Figure6B": ["fig6_fatigue_order_student_data.csv", "fig6B_first_second_score_stats.csv"],
    "Figure6C": ["fig6_fatigue_order_student_data.csv", "fig6C_position_difference_by_sequence.csv"],
    "Figure6D": ["fig6_fatigue_order_student_data.csv", "fig6D_adjusted_fatigue_effect.csv"],
    "Figure6E": ["fig6_fatigue_order_student_data.csv", "fig6E_total_duration_by_sequence.csv"],
}


PANEL_NOTES: dict[str, str] = {
    "Figure2E": (
        "Mixed model: quality_score_5 ~ source*cognitive_level + char_count + "
        "has_vignette + (1|rater_id) + (1|item_id)."
    ),
    "Figure2F": (
        "Consistency panel uses two-way mixed-effects consistency ICC. The plotted "
        "statistic is average-measure ICC(3,k) across four machine runs."
    ),
    "Figure3D": (
        "Figure 3D uses an adjusted response-level binomial model controlling for "
        "block position, training setting/campus, training year, randomized form, "
        "and topic; standard errors are clustered by student."
    ),
    "Figure4B": "Wilson 95% confidence intervals.",
    "Figure4E": "Wilson 95% confidence interval.",
    "Figure5B": "All MAS items are treated as usable: denominator is 50/50.",
    "Figure5E": (
        "Primary AI cost uses the user-provided API total (CNY 28.83); token "
        "estimate is retained as a secondary source-data row."
    ),
}


def panel_key(path: Path) -> str:
    match = re.match(r"(Figure\d+[A-Z])_", path.name)
    if not match:
        raise ValueError(f"Cannot parse panel key from {path.name}")
    return match.group(1)


def workbook_name(key: str) -> str:
    match = re.match(r"Figure(\d+)([A-Z])", key)
    if not match:
        raise ValueError(f"Cannot parse workbook name from {key}")
    return f"source_data_fig_{match.group(1)}_{match.group(2).lower()}.xlsx"


def clean_cell(value):
    if pd.isna(value):
        return None
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating, float)):
        value = float(value)
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    if isinstance(value, (np.bool_, bool)):
        return bool(value)
    text = str(value)
    return re.sub(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]", "", text)


def safe_sheet_name(name: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", name)[:31] or "Sheet"
    candidate = base
    counter = 2
    while candidate in used:
        suffix = f"_{counter}"
        candidate = f"{base[:31-len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def write_df(wb: Workbook, name: str, df: pd.DataFrame, used: set[str]) -> None:
    sheet = wb.create_sheet(safe_sheet_name(name, used))
    if df.empty:
        df = pd.DataFrame({"note": ["No rows"]})
    sheet.append([str(column) for column in df.columns])
    max_rows = min(len(df), 1_048_000)
    for row in df.head(max_rows).itertuples(index=False, name=None):
        sheet.append([clean_cell(value) for value in row])
    sheet.freeze_panes = "A2"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sample_rows = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 200)))
    for col_idx in range(1, sheet.max_column + 1):
        width = 10
        for row in sample_rows:
            value = row[col_idx - 1].value
            if value is not None:
                width = max(width, min(len(str(value)) + 2, 42))
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 200)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def existing_csvs(names: Iterable[str]) -> list[Path]:
    paths = []
    for name in names:
        path = DERIVED / name
        if path.exists():
            paths.append(path)
    return paths


def related_csvs(key: str) -> list[Path]:
    manual = existing_csvs(PANEL_SOURCE_MAP.get(key, []))
    seen: set[Path] = set()
    out: list[Path] = []
    for path in manual:
        resolved = path.resolve()
        if resolved not in seen:
            seen.add(resolved)
            out.append(path)
    return out


def main() -> None:
    manifest_rows = []
    panel_paths = sorted(PANELS.glob("Figure*.pdf"))
    for pdf in panel_paths:
        key = panel_key(pdf)
        csv_paths = related_csvs(key)
        wb = Workbook()
        wb.remove(wb.active)
        used: set[str] = set()
        readme = pd.DataFrame(
            [
                {"field": "panel", "value": key},
                {"field": "figure_file", "value": str(pdf.relative_to(PLOT_ROOT))},
                {"field": "source_data_rule", "value": "Current panel-specific data and derived statistics used by the plotting script."},
                {"field": "note", "value": PANEL_NOTES.get(key, "")},
            ]
        )
        write_df(wb, "README", readme, used)
        provenance = pd.DataFrame(
            [{"type": "figure_pdf", "path": str(pdf.relative_to(PLOT_ROOT))}]
            + [
                {"type": "derived_csv", "path": str(path.relative_to(PLOT_ROOT))}
                for path in csv_paths
            ]
        )
        write_df(wb, "provenance", provenance, used)
        if not csv_paths:
            write_df(
                wb,
                "no_csv_match",
                pd.DataFrame({"note": [f"No derived CSV was mapped for {key}."]}),
                used,
            )
        for path in csv_paths:
            df = pd.read_csv(path, encoding="utf-8-sig")
            write_df(wb, path.stem, df, used)
        out = SOURCE_DIR / workbook_name(key)
        wb.save(out)
        manifest_rows.append(
            {
                "panel": key,
                "figure_file": str(pdf.relative_to(PLOT_ROOT)),
                "source_data_workbook": str(out.relative_to(PLOT_ROOT)),
                "n_derived_csvs": len(csv_paths),
            }
        )
    manifest = pd.DataFrame(manifest_rows)
    manifest.to_csv(SOURCE_DIR / "source_data_manifest.csv", index=False, encoding="utf-8-sig")
    print(f"Wrote {len(manifest_rows)} source-data workbooks to {SOURCE_DIR}")


if __name__ == "__main__":
    main()
