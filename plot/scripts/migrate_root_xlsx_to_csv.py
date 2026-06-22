#!/usr/bin/env python3
"""Export every repository-root XLSX worksheet to CSV, then optionally delete XLSX.

The exported files preserve worksheet rows verbatim (no inferred header), and a
manifest keeps the original workbook/sheet names so plotting code can read the
CSV exports after the source workbooks are removed.
"""
from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
DERIVED = REPO_ROOT / "plot" / "derived_data"
EXPORT_DIR = DERIVED / "source_workbooks"
MANIFEST = DERIVED / "root_xlsx_csv_manifest.csv"


def safe_name(value: str) -> str:
    value = re.sub(r"[^\w.-]+", "_", value.strip(), flags=re.UNICODE)
    return value.strip("_.") or "sheet"


def workbook_slug(path: Path) -> str:
    name = path.stem
    expert = re.search(r"专家\s*(\d+)", name)
    if expert:
        return f"expert_{expert.group(1)}"
    if name == "效率分析":
        return "efficiency_analysis"
    if name == "试卷作答情况":
        return "exam_responses_workbook_1"
    if name == "试卷作答情况 - 2":
        return "exam_responses_workbook_2"
    return safe_name(name)


def export_workbook(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[dict[str, object]] = []
    slug = workbook_slug(path)
    for index, worksheet in enumerate(workbook.worksheets, start=1):
        filename = f"{slug}__{index:02d}__{safe_name(worksheet.title)}.csv"
        output = EXPORT_DIR / filename
        with output.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            for row in worksheet.iter_rows(values_only=True):
                writer.writerow(list(row))
        records.append(
            {
                "workbook_name": path.name,
                "sheet_name": worksheet.title,
                "csv_path": output.relative_to(REPO_ROOT).as_posix(),
                "row_count": worksheet.max_row,
                "column_count": worksheet.max_column,
            }
        )
    workbook.close()
    return records


def verify_exports(records: list[dict[str, object]]) -> None:
    for record in records:
        path = REPO_ROOT / str(record["csv_path"])
        if not path.exists():
            raise FileNotFoundError(f"Missing exported CSV: {path}")
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.reader(handle))
        if len(rows) != int(record["row_count"]):
            raise ValueError(
                f"{path}: expected {record['row_count']} rows, found {len(rows)}"
            )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--delete-originals",
        action="store_true",
        help="Delete the explicitly enumerated repository-root XLSX files after verification.",
    )
    args = parser.parse_args()

    sources = sorted(
        path
        for path in REPO_ROOT.glob("*.xlsx")
        if not path.name.startswith("~$")
    )
    if not sources:
        print("No repository-root XLSX files found; existing CSV exports were left unchanged.")
        return

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    records: list[dict[str, object]] = []
    for source in sources:
        records.extend(export_workbook(source))
    verify_exports(records)

    with MANIFEST.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "workbook_name",
                "sheet_name",
                "csv_path",
                "row_count",
                "column_count",
            ],
        )
        writer.writeheader()
        writer.writerows(records)

    if args.delete_originals:
        for source in sources:
            resolved = source.resolve()
            if resolved.parent != REPO_ROOT.resolve() or resolved.suffix.lower() != ".xlsx":
                raise ValueError(f"Refusing to delete unexpected path: {resolved}")
        for source in sources:
            source.unlink()

    action = "exported and deleted" if args.delete_originals else "exported"
    print(f"{action}: {len(sources)} workbooks, {len(records)} worksheets")
    print(f"Manifest: {MANIFEST}")


if __name__ == "__main__":
    main()
