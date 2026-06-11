#!/usr/bin/env python3
"""Ingest first-author 2026-05-26 supplemental files into agent_readable derived tables.

Inputs:
  agent_readable/data/first_author_update/urology_expert_evaluation_collection/*.xlsx
  agent_readable/data/first_author_update/critical_defect.docx
Outputs:
  agent_readable/derived_data/expert_ratings_updated.csv
  agent_readable/derived_data/expert_rating_item_summary_updated.csv
  agent_readable/derived_data/source_detection_updated.csv
  agent_readable/derived_data/source_detection_metrics_updated.csv
  agent_readable/derived_data/source_task_ratings_updated.csv
  agent_readable/derived_data/critical_defect_taxonomy_updated.csv
"""
from __future__ import annotations
import re
from pathlib import Path
from difflib import SequenceMatcher
import pandas as pd
from openpyxl import load_workbook
from docx import Document

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data" / "first_author_update"
DERIVED = ROOT / "derived_data"
ITEM_MASTER = DERIVED / "item_master.csv"

QG_COMPONENTS = ["qg_fluency", "qg_clarity", "qg_conciseness", "qg_relevance", "qg_consistency", "qg_answerability", "qg_answer_consistency"]
LLM_COMPONENTS = [
    "llm_fluency", "llm_exclusiveness", "llm_explicitness", "llm_goal_alignment",
    "llm_comprehensiveness", "llm_focus", "llm_guess_resistance", "llm_completeness",
    "llm_correctness", "llm_solvability", "llm_absoluteness", "llm_plausibility",
    "llm_reasoning", "llm_feedback", "llm_fairness", "llm_explanation_score",
]

def norm_text(x: object) -> str:
    if x is None:
        return ""
    s = str(x)
    s = re.sub(r"考点还原：.*", "", s, flags=re.S)
    s = re.sub(r"参考答案[:：].*", "", s, flags=re.S)
    s = re.sub(r"[\s\u3000]+", "", s)
    s = re.sub(r"^[0-9一二三四五六七八九十]+[、.．)]", "", s)
    s = re.sub(r"（[A-E]）|\([A-E]\)|【.*?】|（\d+分）", "", s)
    return s[:800]

def best_item_match(qtext: str, candidates: pd.DataFrame) -> tuple[str, int, float]:
    q = norm_text(qtext)
    best = (None, None, -1.0)
    for _, row in candidates.iterrows():
        score = SequenceMatcher(None, q, norm_text(row["question_text"])).ratio()
        if score > best[2]:
            best = (row["item_id"], int(row["paper_item_no"]), float(score))
    return best

def is_data_row(row: tuple) -> bool:
    qtext = row[1] if len(row) > 1 else None
    qg_total = row[10] if len(row) > 10 else None
    return isinstance(qtext, str) and len(qtext.strip()) > 12 and isinstance(qg_total, (int, float))

def clean_rater_id(path: Path) -> str:
    m = re.search(r"专家\s*(\d+)", path.name)
    return f"expert_{m.group(1)}" if m else path.stem[:30]

def parse_yes_no_guess(x: object) -> str | None:
    if x is None:
        return None
    s = str(x).strip()
    if s in {"是", "AI", "ai", "人工智能"}:
        return "MAS"
    if s in {"否", "人工", "Human", "human"}:
        return "Human"
    return None

def ingest_workbooks() -> pd.DataFrame:
    item_master = pd.read_csv(ITEM_MASTER)
    rows = []
    xlsx_files = sorted((DATA_DIR / "urology_expert_evaluation_collection").glob("*.xlsx"))
    for xlsx in xlsx_files:
        wb = load_workbook(xlsx, read_only=True, data_only=True)
        rater_id = clean_rater_id(xlsx)
        for sheet_name, paper, source_true in [("P", "P", "Human"), ("M", "M", "MAS")]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            candidates = item_master[item_master["paper"].eq(paper)].copy()
            for excel_row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
                if not is_data_row(row):
                    continue
                qtext = row[1]
                item_id, paper_item_no, match_score = best_item_match(qtext, candidates)
                values = list(row)
                qg_vals = values[3:10]
                llm_vals = values[11:27]
                qg_total = values[10]
                llm_total = values[27]
                comp_vals = [v for v in qg_vals + llm_vals if isinstance(v, (int, float))]
                out = {
                    "rater_id": rater_id,
                    "source_file": xlsx.name,
                    "sheet": sheet_name,
                    "excel_row": excel_row_idx,
                    "paper": paper,
                    "source_true": source_true,
                    "item_id": item_id,
                    "paper_item_no": paper_item_no,
                    "match_score": match_score,
                    "item_type_from_sheet": values[0],
                    "question_text_from_sheet": qtext,
                    "qgeval_total": qg_total,
                    "qgeval_score_5": qg_total / 7 if isinstance(qg_total, (int, float)) else None,
                    "llm_total": llm_total,
                    "llm_score_5": llm_total / 16 if isinstance(llm_total, (int, float)) else None,
                    "quality_score_5": sum(comp_vals) / len(comp_vals) if comp_vals else None,
                    "source_guess": parse_yes_no_guess(values[28] if len(values) > 28 else None),
                }
                for name, val in zip(QG_COMPONENTS, qg_vals):
                    out[name] = val
                for name, val in zip(LLM_COMPONENTS, llm_vals):
                    out[name] = val
                rows.append(out)
    df = pd.DataFrame(rows)
    return df

def ingest_critical_defect_taxonomy() -> pd.DataFrame:
    docx = DATA_DIR / "critical_defect.docx"
    if not docx.exists():
        return pd.DataFrame(columns=["defect_category", "source_file", "note"])
    doc = Document(docx)
    cats = []
    for p in doc.paragraphs:
        text = p.text.strip().strip("；;")
        if not text:
            continue
        if "每个排除项" in text:
            note = text
            continue
        cats.append(text)
    return pd.DataFrame({"defect_category": cats, "source_file": docx.name, "note": "categories requested for item-flow exclusion reporting"})

def source_metrics(src: pd.DataFrame) -> pd.DataFrame:
    d = src.dropna(subset=["source_guess"]).copy()
    if d.empty:
        return pd.DataFrame()
    d["correct_source_guess"] = (d["source_guess"] == d["source_true"]).astype(int)
    accuracy = d["correct_source_guess"].mean()
    sens_mas = d.loc[d["source_true"].eq("MAS"), "correct_source_guess"].mean()
    spec_human = d.loc[d["source_true"].eq("Human"), "correct_source_guess"].mean()
    bal = (sens_mas + spec_human) / 2
    rows = [
        {"metric": "accuracy", "estimate": accuracy, "numerator": int(d["correct_source_guess"].sum()), "denominator": int(len(d)), "inference_note": "first_author_update_expert_judgments"},
        {"metric": "balanced_accuracy", "estimate": bal, "numerator": None, "denominator": int(len(d)), "inference_note": "mean_of_MAS_sensitivity_and_Human_specificity"},
        {"metric": "sensitivity_mas", "estimate": sens_mas, "numerator": int(d.loc[d["source_true"].eq("MAS"), "correct_source_guess"].sum()), "denominator": int(d["source_true"].eq("MAS").sum()), "inference_note": "correctly_identified_MAS"},
        {"metric": "specificity_human", "estimate": spec_human, "numerator": int(d.loc[d["source_true"].eq("Human"), "correct_source_guess"].sum()), "denominator": int(d["source_true"].eq("Human").sum()), "inference_note": "correctly_identified_Human"},
    ]
    return pd.DataFrame(rows)

def main() -> None:
    DERIVED.mkdir(exist_ok=True)
    ratings = ingest_workbooks()
    ratings.to_csv(DERIVED / "expert_ratings_updated.csv", index=False)
    item_summary = ratings.groupby(["item_id", "source_true"], as_index=False).agg(
        n_expert_ratings=("quality_score_5", "count"),
        expert_quality_score_5_mean=("quality_score_5", "mean"),
        expert_quality_score_5_sd=("quality_score_5", "std"),
        expert_qgeval_score_5_mean=("qgeval_score_5", "mean"),
        expert_llm_score_5_mean=("llm_score_5", "mean"),
        mean_match_score=("match_score", "mean"),
    )
    item_summary.to_csv(DERIVED / "expert_rating_item_summary_updated.csv", index=False)

    src = ratings.dropna(subset=["source_guess"])[["rater_id", "item_id", "source_true", "source_guess", "quality_score_5"]].copy()
    src["correct_source_guess"] = (src["source_true"] == src["source_guess"]).astype(int)
    src.to_csv(DERIVED / "source_detection_updated.csv", index=False)
    source_metrics(src).to_csv(DERIVED / "source_detection_metrics_updated.csv", index=False)
    src.rename(columns={"quality_score_5": "source_task_rating_5"}).to_csv(DERIVED / "source_task_ratings_updated.csv", index=False)
    ingest_critical_defect_taxonomy().to_csv(DERIVED / "critical_defect_taxonomy_updated.csv", index=False)
    print(f"Wrote {len(ratings)} expert rating rows from update.")
    print(ratings.groupby(["source_true"]).size())
    print("Source judgments:", len(src))
    print(source_metrics(src).to_string(index=False))
    low_match = ratings[ratings["match_score"] < 0.55]
    low_match[["rater_id", "source_file", "sheet", "excel_row", "item_id", "paper_item_no", "match_score", "question_text_from_sheet"]].to_csv(DERIVED / "expert_rating_match_qc_updated.csv", index=False)
    if not low_match.empty:
        print("WARNING: low fuzzy matches detected; see derived_data/expert_rating_match_qc_updated.csv")
        print(low_match[["rater_id", "sheet", "excel_row", "item_id", "match_score"]].head(20).to_string(index=False))

if __name__ == "__main__":
    main()
