#!/usr/bin/env python3
"""Legacy source-data exporter retained only for historical figure-numbering review.

Each workbook contains: README, provenance, raw values used by the panel, intermediate
calculation tables, and final plot data. Excel creation is handled by artifact_tool.
"""
from __future__ import annotations
from pathlib import Path
import math, json, re, shutil, zipfile
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[2]
ROOT = REPO_ROOT / "plot"
DERIVED = ROOT / "data" / "derived"
DATA = ROOT / "data"
FIGS = REPO_ROOT / "outputs" / "figures" / "panels"
SOURCE_DIR = REPO_ROOT / "outputs" / "figure_source_data_legacy"
DOCS = ROOT / "docs"
SOURCE_DIR.mkdir(parents=True, exist_ok=True)

LEVELS = ["knowledge", "application", "reasoning"]

def read_csv(name: str) -> pd.DataFrame:
    return pd.read_csv(DERIVED / name)

def clean_value(x):
    if pd.isna(x):
        return None
    if isinstance(x, (np.integer,)):
        return int(x)
    if isinstance(x, (np.floating,)):
        if math.isnan(float(x)) or math.isinf(float(x)):
            return None
        return float(x)
    if isinstance(x, (np.bool_,)):
        return bool(x)
    return x

def df_to_matrix(df: pd.DataFrame, max_rows: int | None = None):
    d = df.copy()
    if max_rows is not None:
        d = d.head(max_rows)
    cols = [str(c) for c in d.columns]
    rows = [[clean_value(v) for v in row] for row in d.itertuples(index=False, name=None)]
    return [cols] + rows

def safe_sheet_name(name: str) -> str:
    s = re.sub(r"[\\/*?:\[\]]", "_", name)[:31]
    return s or "Sheet"

def write_sheet(wb, name: str, df: pd.DataFrame, width: int = 18):
    sh = wb.create_sheet(safe_sheet_name(name))
    mat = df_to_matrix(df)
    if not mat:
        mat = [["empty"]]
    for row in mat:
        sh.append(row)
    sh.freeze_panes = "A2"
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sh[1]:
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    max_col = len(mat[0]) if mat else 1
    sample_rows = mat[:200]
    for c in range(1, max_col + 1):
        max_len = 0
        for row in sample_rows:
            if c <= len(row) and row[c-1] is not None:
                max_len = max(max_len, min(len(str(row[c-1])), 80))
        sh.column_dimensions[get_column_letter(c)].width = min(max(max_len + 2, 10), 36)
    for row in sh.iter_rows(min_row=1, max_row=min(sh.max_row, 200), max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    return sh

def write_workbook(filename: str, panel: str, description: str, sheets: dict[str, pd.DataFrame], source_files: list[str], figure_files: list[str], notes: list[str] | None = None):
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)
    readme = pd.DataFrame({
        "field": ["panel", "description", "created_by", "content_rule", "notes"],
        "value": [panel, description, "generate_panel_source_data_and_review.py", "Includes raw values used for plotting and intermediate/derived values used by the plotting code.", " | ".join(notes or [])]
    })
    write_sheet(wb, "README", readme, 36)
    prov = pd.DataFrame({
        "type": ["source_file"]*len(source_files) + ["figure_file"]*len(figure_files),
        "path": source_files + figure_files
    })
    write_sheet(wb, "provenance", prov, 36)
    for name, df in sheets.items():
        write_sheet(wb, name, df)
    out = SOURCE_DIR / filename
    wb.save(out)
    return out

def diff_summary(df, value_col, group_col="source_true"):
    g = df.groupby(group_col)[value_col].agg(["count", "mean", "std"]).reset_index()
    vals = {r[group_col]: r["mean"] for _, r in g.iterrows()}
    out = g.copy()
    if "MAS" in vals and "Human" in vals:
        diff = vals["MAS"] - vals["Human"]
        out = pd.concat([out, pd.DataFrame([{group_col:"MAS_minus_Human", "count":None, "mean":diff, "std":None}])], ignore_index=True)
    return out

def bootstrap_diff(raw, value_col, by_col=None, reps=200, seed=1):
    rng = np.random.default_rng(seed)
    rows=[]; reps_rows=[]
    groups = [("overall", raw)] if by_col is None else list(raw.groupby(by_col))
    for g, sub in groups:
        a = sub.loc[sub.source_true.eq("MAS"), value_col].dropna().to_numpy(float)
        h = sub.loc[sub.source_true.eq("Human"), value_col].dropna().to_numpy(float)
        if len(a)==0 or len(h)==0:
            continue
        vals=[]
        for i in range(reps):
            est = rng.choice(a, len(a), True).mean() - rng.choice(h, len(h), True).mean()
            vals.append(est)
            reps_rows.append({"group":g, "bootstrap_replicate":i+1, "mas_minus_human":est})
        rows.append({"group":g, "estimate":a.mean()-h.mean(), "ci_low":np.percentile(vals, 2.5), "ci_high":np.percentile(vals, 97.5), "n_mas":len(a), "n_human":len(h), "seed":seed, "reps":reps})
    return pd.DataFrame(rows), pd.DataFrame(reps_rows)

def balanced_accuracy_boot(src, reps=500, seed=7):
    src=src.dropna(subset=["source_guess"]).copy()
    src["correct_source_guess"]=(src.source_guess==src.source_true).astype(int)
    mas = src[src.source_true.eq("MAS")]["correct_source_guess"].to_numpy()
    hum = src[src.source_true.eq("Human")]["correct_source_guess"].to_numpy()
    rng=np.random.default_rng(seed)
    rows=[]
    for i in range(reps):
        rows.append({"bootstrap_replicate": i+1, "balanced_accuracy": (rng.choice(mas, len(mas), True).mean()+rng.choice(hum, len(hum), True).mean())/2})
    reps_df=pd.DataFrame(rows)
    summary=pd.DataFrame([{"metric":"balanced_accuracy", "estimate":(mas.mean()+hum.mean())/2, "ci_low":reps_df.balanced_accuracy.quantile(.025), "ci_high":reps_df.balanced_accuracy.quantile(.975), "n_judgments":len(src), "n_mas_true":len(mas), "n_human_true":len(hum), "seed":seed, "reps":reps}])
    return summary, reps_df

def main():
    manifest=[]
    # Shared tables
    item = read_csv("item_master.csv")
    expert = read_csv("expert_ratings_updated.csv")
    expert_item = expert.merge(item[["item_id","cognitive_level","topic"]], on="item_id", how="left")
    responses = read_csv("responses.csv")
    assign = read_csv("exam_form_assignment.csv")
    block = read_csv("block_scores.csv")
    paired = read_csv("paired_block_scores.csv")
    defect = read_csv("defect_adjudication_proxy.csv")
    ctt = read_csv("ctt_item_analysis.csv")
    safety = read_csv("machine_safety_screening_source_summary.csv")
    safety_run = read_csv("machine_safety_screening_by_run.csv")
    mas_time = read_csv("mas_question_generation_time.csv")
    table2 = read_csv("table2_examinee_baseline_randomization_balance.csv")
    table3 = read_csv("table3_primary_key_secondary_endpoints.csv")
    workflow_template = read_csv("workflow_total_time_cost_TEMPLATE.csv")
    sourceupdate = read_csv("source_detection_updated.csv")
    source_task = read_csv("source_task_ratings_updated.csv")
    crit_tax = read_csv("critical_defect_taxonomy_updated.csv")

    def add(fname, panel, desc, sheets, sources, figs, notes=None):
        out=write_workbook(fname, panel, desc, sheets, sources, figs, notes)
        manifest.append({"panel":panel,"file":str(out.relative_to(ROOT)),"description":desc,"n_sheets":len(sheets)+2})

    # Figure 1
    add("source_data_fig_1_a.xlsx","Figure 1A","Inputs and MAS generation workflow counts and named components.",{
        "plot_components": pd.DataFrame([
            {"component":"Human expert bank", "display_value":"n=775", "raw_or_source":"Codex/project description and figure prompt"},
            {"component":"Authorized source materials", "display_value":"CUA/EAU guidelines; textbook; syllabus", "raw_or_source":"plot requirements"},
            {"component":"Exam blueprint constraints", "display_value":"A1/A2/A3-A4/B/X; topics; cognitive levels", "raw_or_source":"exam_blueprint.md + item_master.csv"},
            {"component":"MAS candidate bank", "display_value":"n=3,676", "raw_or_source":"gpt_image2_figure1_prompt.md"},
        ]),
        "item_master_summary": item.groupby(["source_true","item_type","cognitive_level"]).size().reset_index(name="n_items"),
        "mas_generation_time_raw": mas_time,
    }, ["derived_data/item_master.csv","derived_data/mas_question_generation_time.csv","docs/gpt_image2_figure1_prompt.md"], ["figures/panels/Figure1A_workflow_inputs_v3.png","figures/panels/Figure1A_workflow_inputs_v3.pdf"])
    add("source_data_fig_1_b.xlsx","Figure 1B","AI safety gate domains and machine screening source summaries.",{
        "gate_checklist": pd.DataFrame({"check":["Answer-key legality","Guideline consistency","Single best answer","Stem and option clarity","Major / critical defect flag"],"role":["QC/proxy only"]*5}),
        "screening_source_summary": safety,
        "screening_run_raw": safety_run,
    }, ["derived_data/machine_safety_screening_source_summary.csv","derived_data/machine_safety_screening_by_run.csv"], ["figures/panels/Figure1B_safety_gate.pdf"])
    add("source_data_fig_1_c.xlsx","Figure 1C","Expert intake review boundary and lock-state data used in workflow panel.",{
        "workflow_boundary": pd.DataFrame([
            {"step":"Parallel expert intake review", "purpose":"Safety and assembly control before administration", "included_in_final_blind_endpoint":"No"},
            {"step":"Source key locked", "purpose":"Prevent source disclosure until blind rating complete", "included_in_final_blind_endpoint":"Boundary condition"},
        ]),
        "critical_defect_taxonomy": crit_tax,
    }, ["derived_data/critical_defect_taxonomy_updated.csv","data/first_author_update/critical_defect.docx"], ["figures/panels/Figure1C_two_sequence_order.pdf"])
    add("source_data_fig_1_d.xlsx","Figure 1D","Randomized two-sequence exam design and form/setting assignment data.",{
        "form_assignment_raw": assign,
        "form_summary": table2,
        "order_schema": pd.DataFrame([{"form":"A","order":"Human -> MAS"},{"form":"B","order":"MAS -> Human"}]),
    }, ["derived_data/exam_form_assignment.csv","derived_data/table2_examinee_baseline_randomization_balance.csv"], ["figures/panels/Figure1D_training_setting.pdf"])
    add("source_data_fig_1_e.xlsx","Figure 1E","Endpoint-domain tiles shown in the design figure.",{
        "endpoint_domains": table3,
        "domain_tiles": pd.DataFrame({"tile":["Expert quality noninferiority","Major / critical defects","Cognitive-level boundaries","Source detectability","Student performance","Aggregate quality-adjusted efficiency"]}),
    }, ["derived_data/table3_primary_key_secondary_endpoints.csv"], ["figures/panels/Figure1E_endpoint_domains.pdf"])

    # Figure 2
    summ2a, boot2a = bootstrap_diff(expert, "quality_score_5", reps=500, seed=21)
    add("source_data_fig_2_a.xlsx","Figure 2A","Primary expert-rated composite quality difference, MAS minus Human.",{
        "raw_expert_ratings": expert,
        "group_means": diff_summary(expert,"quality_score_5"),
        "plot_data": summ2a,
        "bootstrap_replicates": boot2a,
    }, ["derived_data/expert_ratings_updated.csv"], ["figures/panels/Figure2A_quality_difference.pdf"], ["update expert ratings are used for the formal source-data file."])
    dim_cols=[c for c in expert.columns if c.startswith("qg_") or c.startswith("ulm_")]
    dim_rows=[]
    for c in dim_cols:
        for src, sub in expert.groupby("source_true"):
            dim_rows.append({"dimension":c,"source_true":src,"n":sub[c].notna().sum(),"mean":sub[c].mean(),"sd":sub[c].std()})
        vals=expert.groupby("source_true")[c].mean().to_dict()
        if "MAS" in vals and "Human" in vals:
            dim_rows.append({"dimension":c,"source_true":"MAS_minus_Human","n":None,"mean":vals["MAS"]-vals["Human"],"sd":None})
    add("source_data_fig_2_b.xlsx","Figure 2B","Expert rating component/domain scores and intermediate source differences.",{
        "raw_dimension_scores": expert[["rater_id","item_id","source_true"]+dim_cols],
        "dimension_summary": pd.DataFrame(dim_rows),
    }, ["derived_data/expert_ratings_updated.csv"], ["figures/panels/Figure2B_dimension_scores.pdf"])
    defect2=defect.copy(); defect2["any_major_or_critical_defect"]=((defect2.final_major_defect.fillna(0)>0)|(defect2.final_critical_defect.fillna(0)>0)).astype(int)
    add("source_data_fig_2_c.xlsx","Figure 2C","Major/critical defect flags and source-level defect rates.",{
        "raw_defect_proxy": defect2,
        "defect_rate_summary": defect2.groupby("source_true").agg(n_items=("item_id","count"), major_rate=("final_major_defect","mean"), critical_rate=("final_critical_defect","mean"), any_major_or_critical_rate=("any_major_or_critical_defect","mean")).reset_index(),
        "critical_defect_taxonomy": crit_tax,
    }, ["derived_data/defect_adjudication_proxy.csv","derived_data/critical_defect_taxonomy_updated.csv"], ["figures/panels/Figure2C_defect_flags.pdf"], ["Formal item-level expert defect adjudication was not present in the update packet; this uses the existing adjudication proxy and taxonomy."])
    add("source_data_fig_2_d.xlsx","Figure 2D","Defect adjudication workflow and category definitions.",{
        "workflow_steps": pd.DataFrame([
            {"step":1,"label":"Machine QC/proxy flags","data":"machine_safety_screening_by_run.csv"},
            {"step":2,"label":"Human expert review / adjudication required","data":"defect_adjudication_proxy.csv currently available"},
            {"step":3,"label":"Critical-defect taxonomy","data":"critical_defect_taxonomy_updated.csv"},
        ]),
        "defect_proxy_raw": defect2,
        "critical_defect_taxonomy": crit_tax,
    }, ["derived_data/defect_adjudication_proxy.csv","derived_data/critical_defect_taxonomy_updated.csv"], ["figures/panels/Figure2D_defect_workflow.pdf"])
    add("source_data_fig_2_e.xlsx","Figure 2E","Expert rating consistency intermediate summaries.",{
        "raw_expert_ratings": expert[["rater_id","item_id","source_true","quality_score_5","qgeval_score_5","ulm_score_5"]],
        "rater_summary": expert.groupby(["rater_id","source_true"]).agg(n_ratings=("item_id","count"), mean_quality=("quality_score_5","mean"), sd_quality=("quality_score_5","std")).reset_index(),
        "item_rating_summary": expert.groupby(["item_id","source_true"]).agg(n_raters=("rater_id","nunique"), mean_quality=("quality_score_5","mean"), sd_quality=("quality_score_5","std")).reset_index(),
    }, ["derived_data/expert_ratings_updated.csv"], ["figures/panels/Figure2E_run_consistency.pdf"])

    # Figure 3
    q_raw=expert_item.dropna(subset=["cognitive_level"])
    q_summ, q_boot=bootstrap_diff(q_raw,"quality_score_5",by_col="cognitive_level",reps=200,seed=3)
    plot3a=read_csv("fig3A_quality_by_cognitive_level_updated.csv")
    add("source_data_fig_3_a.xlsx","Figure 3A","Expert-rated quality difference by cognitive level.",{
        "raw_expert_ratings_with_level": q_raw,
        "intermediate_group_means": q_raw.groupby(["cognitive_level","source_true"]).agg(n=("quality_score_5","count"), mean=("quality_score_5","mean"), sd=("quality_score_5","std")).reset_index(),
        "bootstrap_replicates": q_boot,
        "plot_data": plot3a,
    }, ["derived_data/expert_ratings_updated.csv","derived_data/item_master.csv","derived_data/fig3A_quality_by_cognitive_level_updated.csv"], ["figures/panels/revised/Figure3A_quality_by_cognitive_level.png","figures/panels/revised/Figure3A_quality_by_cognitive_level.pdf"])
    d_raw=defect2.merge(item[["item_id","cognitive_level"]],on="item_id",how="left")
    d_raw["any_major_or_critical_defect_pp"]=d_raw["any_major_or_critical_defect"]*100
    d_summ, d_boot=bootstrap_diff(d_raw,"any_major_or_critical_defect_pp",by_col="cognitive_level",reps=200,seed=4)
    add("source_data_fig_3_b.xlsx","Figure 3B","Item-writing defect-risk difference by cognitive level.",{
        "raw_defect_with_level": d_raw,
        "intermediate_group_means": d_raw.groupby(["cognitive_level","source_true"]).agg(n=("item_id","count"), mean_defect_pp=("any_major_or_critical_defect_pp","mean")).reset_index(),
        "bootstrap_replicates": d_boot,
        "plot_data": read_csv("fig3B_defect_risk_by_cognitive_level_updated.csv"),
    }, ["derived_data/defect_adjudication_proxy.csv","derived_data/item_master.csv","derived_data/fig3B_defect_risk_by_cognitive_level_updated.csv"], ["figures/panels/revised/Figure3B_defect_risk_by_cognitive_level.png","figures/panels/revised/Figure3B_defect_risk_by_cognitive_level.pdf"])
    r_raw=responses.merge(item[["item_id","cognitive_level"]],on="item_id",how="left")
    r_raw["correct_pp"]=r_raw["correct"].astype(float)*100
    r_summ, r_boot=bootstrap_diff(r_raw,"correct_pp",by_col="cognitive_level",reps=200,seed=5)
    add("source_data_fig_3_c.xlsx","Figure 3C","Student correct-answer rate difference by cognitive level.",{
        "raw_responses_with_level": r_raw,
        "intermediate_group_means": r_raw.groupby(["cognitive_level","source_true"]).agg(n_responses=("correct","count"), correct_rate=("correct","mean"), correct_pp=("correct_pp","mean")).reset_index(),
        "bootstrap_replicates": r_boot,
        "plot_data": read_csv("fig3C_student_accuracy_by_cognitive_level_updated.csv"),
    }, ["derived_data/responses.csv","derived_data/item_master.csv","derived_data/fig3C_student_accuracy_by_cognitive_level_updated.csv"], ["figures/panels/revised/Figure3C_student_accuracy_by_cognitive_level.png","figures/panels/revised/Figure3C_student_accuracy_by_cognitive_level.pdf"])
    inter=q_raw.groupby(["cognitive_level","source_true"]).agg(n=("quality_score_5","count"), mean_quality=("quality_score_5","mean"), sd_quality=("quality_score_5","std")).reset_index()
    add("source_data_fig_3_d.xlsx","Figure 3D","Source by cognitive-level interaction descriptive data.",{
        "raw_expert_ratings_with_level": q_raw,
        "interaction_cell_means": inter,
        "source_difference_by_level": q_summ,
    }, ["derived_data/expert_ratings_updated.csv","derived_data/item_master.csv"], ["figures/panels/Figure3D_source_cognitive_interaction.pdf"], ["No revised update panel file was present for Figure 3D; source data supports the existing panel concept."])
    ctt_raw=ctt.merge(item[["item_id","cognitive_level"]],on="item_id",how="left")
    add("source_data_fig_3_e.xlsx","Figure 3E","Classical test theory item statistics by cognitive level.",{
        "raw_ctt_with_level": ctt_raw,
        "ctt_summary": ctt_raw.groupby(["cognitive_level","source_true"]).agg(n_items=("item_id","count"), mean_difficulty=("difficulty","mean"), mean_discrimination=("discrimination","mean"), sd_difficulty=("difficulty","std"), sd_discrimination=("discrimination","std")).reset_index(),
    }, ["derived_data/ctt_item_analysis.csv","derived_data/item_master.csv"], ["figures/panels/Figure3E_ctt_by_cognitive_level.pdf"])

    # Figure 4
    ba, ba_boot=balanced_accuracy_boot(sourceupdate, reps=500, seed=7)
    add("source_data_fig_4_a.xlsx","Figure 4A","Balanced source-identification accuracy under blinded conditions.",{
        "raw_source_judgments": sourceupdate,
        "intermediate_by_true_source": sourceupdate.dropna(subset=["source_guess"]).assign(correct_source_guess=lambda d:(d.source_guess==d.source_true).astype(int)).groupby("source_true").agg(n=("correct_source_guess","count"), accuracy=("correct_source_guess","mean")).reset_index(),
        "bootstrap_replicates": ba_boot,
        "plot_data": read_csv("fig4A_source_detection_accuracy_updated.csv"),
    }, ["derived_data/source_detection_updated.csv","derived_data/fig4A_source_detection_accuracy_updated.csv"], ["figures/panels/revised/Figure4A_source_detection_accuracy.png","figures/panels/revised/Figure4A_source_detection_accuracy.pdf"])
    src=sourceupdate.dropna(subset=["source_guess"]).copy()
    counts=pd.crosstab(src.source_true,src.source_guess).reindex(index=["Human","MAS"],columns=["Human","MAS"],fill_value=0).reset_index()
    prop=counts.copy(); prop[["Human","MAS"]]=prop[["Human","MAS"]].div(prop[["Human","MAS"]].sum(axis=1),axis=0)*100
    add("source_data_fig_4_b.xlsx","Figure 4B","True source by guessed source confusion matrix and stacked percentages.",{
        "raw_source_judgments": src,
        "confusion_counts": counts,
        "confusion_percent": prop,
        "plot_data": read_csv("fig4B_source_judgment_confusion_matrix_updated.csv"),
    }, ["derived_data/source_detection_updated.csv","derived_data/fig4B_source_judgment_confusion_matrix_updated.csv"], ["figures/panels/revised/Figure4B_source_judgment_confusion_matrix.png","figures/panels/revised/Figure4B_source_judgment_confusion_matrix.pdf"])
    add("source_data_fig_4_c.xlsx","Figure 4C","Ratings collected during the source-identification task.",{
        "raw_source_task_ratings": source_task,
        "boxplot_summary": source_task.dropna(subset=["source_task_rating_5"]).groupby("source_true")["source_task_rating_5"].agg(n="count", mean="mean", sd="std", median="median", q1=lambda x:x.quantile(.25), q3=lambda x:x.quantile(.75), min="min", max="max").reset_index(),
    }, ["derived_data/source_task_ratings_updated.csv"], ["figures/panels/revised/Figure4C_source_task_ratings.png","figures/panels/revised/Figure4C_source_task_ratings.pdf"])
    add("source_data_fig_4_d.xlsx","Figure 4D","Aggregate workflow total time source-data template and available MAS timing data.",{
        "manual_time_template": workflow_template,
        "mas_generation_time_raw": mas_time,
        "status": pd.DataFrame([{"requirement":"Total human and MAS workflow time", "status":"Partially unmet", "reason":"Human/manual total workflow time has not been populated in workflow_total_time_cost_TEMPLATE.csv."}]),
    }, ["derived_data/workflow_total_time_cost_TEMPLATE.csv","derived_data/mas_question_generation_time.csv"], ["figures/panels/Figure4D_workflow_total_time.pdf"])
    add("source_data_fig_4_e.xlsx","Figure 4E","Quality-adjusted time calculation template.",{
        "manual_time_template": workflow_template,
        "nondefective_counts_from_proxy": defect2.groupby("source_true").agg(number_final_items=("item_id","count"), number_nondefective_items=("any_major_or_critical_defect", lambda x:int((x==0).sum()))).reset_index(),
        "calculation_steps": pd.DataFrame([
            {"step":1,"formula":"time_per_final_item = total_minutes / number_final_items"},
            {"step":2,"formula":"time_per_nondefective_item = total_minutes / number_nondefective_items"},
            {"step":3,"formula":"quality-adjusted time ratio = MAS time_per_nondefective_item / Human time_per_nondefective_item"},
        ]),
    }, ["derived_data/workflow_total_time_cost_TEMPLATE.csv","derived_data/defect_adjudication_proxy.csv"], ["figures/panels/Figure4E_quality_adjusted_time.pdf"], ["Time inputs are still a template, so final numerical plot data cannot be confirmed."])
    add("source_data_fig_4_f.xlsx","Figure 4F","Efficiency sensitivity-analysis parameter grid template.",{
        "manual_time_template": workflow_template,
        "sensitivity_grid_template": pd.DataFrame([{ "human_total_minutes_multiplier":m, "mas_total_minutes_multiplier":n, "calculation":"recompute quality-adjusted time ratio after applying multipliers"} for m in [0.5,1.0,1.5,2.0] for n in [0.5,1.0,1.5,2.0]]),
        "status": pd.DataFrame([{"requirement":"Sensitivity analysis for aggregate efficiency", "status":"Not fully computable", "reason":"Base total human/manual time not provided."}]),
    }, ["derived_data/workflow_total_time_cost_TEMPLATE.csv"], ["figures/panels/Figure4F_efficiency_sensitivity.pdf"])

    # Figure 5
    add("source_data_fig_5_a.xlsx","Figure 5A","Two-sequence block-order schema and participant assignment data.",{
        "order_schema": pd.DataFrame([{"form":"A","first_block":"Human","second_block":"MAS"},{"form":"B","first_block":"MAS","second_block":"Human"}]),
        "form_assignment_raw": assign,
        "block_scores_raw": block,
    }, ["derived_data/exam_form_assignment.csv","derived_data/block_scores.csv"], ["figures/panels/Figure5A_order_schema.pdf"])
    wide=block.pivot_table(index=["student_id","form","training_setting"],columns="source_true",values="score_percent").reset_index()
    wide["mas_minus_human"]=wide["MAS"]-wide["Human"]
    long=wide.melt(id_vars=["student_id","form","training_setting","mas_minus_human"], value_vars=["Human","MAS"], var_name="source_true", value_name="score_percent")
    add("source_data_fig_5_b.xlsx","Figure 5B","Paired block scores for each participant, colored by sequence group.",{
        "raw_block_scores": block,
        "wide_paired_scores": wide,
        "plot_long_format": long,
        "form_summary": wide.groupby("form").agg(n_students=("student_id","count"), mean_human=("Human","mean"), mean_mas=("MAS","mean"), mean_mas_minus_human=("mas_minus_human","mean")).reset_index(),
    }, ["derived_data/block_scores.csv","derived_data/fig5_block_score_differences_updated.csv"], ["figures/panels/revised/Figure5B_paired_block_scores.png","figures/panels/revised/Figure5B_paired_block_scores.pdf"])
    add("source_data_fig_5_c.xlsx","Figure 5C","Within-participant MAS-minus-Human score differences by randomized sequence.",{
        "wide_paired_scores": wide,
        "boxplot_summary_by_form": wide.groupby("form")["mas_minus_human"].agg(n="count", mean="mean", sd="std", median="median", q1=lambda x:x.quantile(.25), q3=lambda x:x.quantile(.75), min="min", max="max").reset_index(),
        "plot_data": read_csv("fig5_block_score_differences_updated.csv"),
    }, ["derived_data/block_scores.csv","derived_data/fig5_block_score_differences_updated.csv"], ["figures/panels/revised/Figure5C_individual_differences_by_sequence.png","figures/panels/revised/Figure5C_individual_differences_by_sequence.pdf"])
    resp5=responses.merge(item[["item_id","topic","cognitive_level"]],on="item_id",how="left").merge(assign[["student_id","form","order_group","training_setting","training_year"]],on="student_id",how="left",suffixes=("","_assign"))
    add("source_data_fig_5_d.xlsx","Figure 5D","Adjusted correct-answer rate difference, MAS minus Human, by sequence.",{
        "raw_model_input": resp5,
        "model_output_plot_data": read_csv("fig5D_correct_rate_difference_by_sequence_updated.csv"),
        "model_formula": pd.DataFrame([{"formula":"participant-level bootstrap of MAS-Human correct-rate differences; within-sequence source and block position are aliased, so fully adjusted source coefficients are not identifiable in this dataset"}]),
    }, ["derived_data/responses.csv","derived_data/item_master.csv","derived_data/exam_form_assignment.csv","derived_data/fig5D_correct_rate_difference_by_sequence_updated.csv"], ["figures/panels/revised/Figure5D_correct_rate_difference_by_sequence.png","figures/panels/revised/Figure5D_correct_rate_difference_by_sequence.pdf"])
    add("source_data_fig_5_e.xlsx","Figure 5E","Exploratory training-setting stratified MAS-minus-Human block score differences.",{
        "wide_paired_scores": wide,
        "boxplot_summary_by_setting": wide.groupby("training_setting")["mas_minus_human"].agg(n="count", mean="mean", sd="std", median="median", q1=lambda x:x.quantile(.25), q3=lambda x:x.quantile(.75), min="min", max="max").reset_index(),
    }, ["derived_data/block_scores.csv","derived_data/fig5_block_score_differences_updated.csv"], ["figures/panels/revised/Figure5E_setting_stratified_differences.png","figures/panels/revised/Figure5E_setting_stratified_differences.pdf"], ["Exploratory only; not a confirmatory between-setting heterogeneity test."])

    manifest_df=pd.DataFrame(manifest)
    manifest_df.to_csv(SOURCE_DIR/"source_data_manifest.csv",index=False)

    review = f"""# update plus source-data audit report

Generated by `scripts/generate_panel_source_data_and_review.py`.

## Reviewer source-data files

Legacy-numbered figure panels have corresponding Excel workbooks in `outputs/figure_source_data_legacy/`. Current reviewer workbooks are generated by `scripts/plotting/export_panel_source_data.py` into `outputs/figure_source_data/`.

- `README`: panel identity and notes.
- `provenance`: input data files and corresponding figure files.
- raw values used by that panel.
- intermediate calculation tables, including group summaries, source differences, percent conversions, and bootstrap replicate tables where the plotting code used bootstrap intervals.
- final plot data where a derived plot-data CSV already exists.

A machine-readable index is saved as `source_data_manifest.csv`.

## Global review of first-author requirements

### Completed

1. **update expert-rating data incorporated.** Expert workbooks from `data/first_author_update/urology_expert_evaluation_collection/` were parsed into `derived_data/expert_ratings_updated.csv`, `expert_rating_item_summary_updated.csv`, `source_detection_updated.csv`, `source_task_ratings_updated.csv`, and QC tables.
2. **Figure 3A-C revised to use the update expert data where applicable.** Figure 3A now uses expert quality ratings by cognitive level; Figure 3B keeps defect-risk analysis by cognitive level using the available defect proxy; Figure 3C uses response-level accuracy by cognitive level.
3. **Figure 4A-C revised per the update recommendation.** Figure 4A reports balanced source-identification accuracy with a chance line and near-chance band; Figure 4B is true-source by guessed-source structure rather than only correct/incorrect counts; Figure 4C uses source-task ratings with a clearer interpretation boundary.
4. **Figure 5B-E revised per the fatigue/order-effect recommendations.** The revised panels show Form A/B ordering, MAS-Human direction, zero reference where appropriate, adjusted/bootstrapped difference output, and exploratory setting-stratified interpretation.
5. **Per-panel source-data workbooks generated.** All Figure 1A-E, 2A-E, 3A-E, 4A-F, and 5A-E panels now have source-data Excel files.

### Still partially unmet or requiring author input

1. **Formal item-level major/critical defect adjudication remains incomplete.** The update `critical_defect.docx` supplies taxonomy and flow-diagram requirements, but not a complete per-item expert defect adjudication table. Therefore Figure 2C/2D and Figure 3B source-data files transparently use `defect_adjudication_proxy.csv` plus the update taxonomy.
2. **Aggregate efficiency panels Figure 4D-F remain limited by missing human/manual workflow time.** `workflow_total_time_cost_TEMPLATE.csv` is still empty. Source-data files for Figure 4D-F include the template, available MAS timing, and formulas/sensitivity-grid templates, but final numerical efficiency comparisons cannot be considered complete until total human/manual time and nondefective-item denominators are populated.
3. **Figure 1 is still represented as separate panel PDFs/PNGs rather than one fully integrated final multi-panel artwork.** Source-data files exist for Figure 1A-E, and a gpt-image2 prompt exists, but a final journal-ready integrated Figure 1 image may still require graphic assembly or manual design review.
4. **Figure 3D and Figure 3E were not specifically updated by the update package.** Source-data files were generated from the current expert-rating, item-master, and CTT tables, but no revised update panel images were present for these two panels.
5. **Figure 2 final blind-defect and expert-consistency statistics may need methodological confirmation.** update expert scoring enables formal source-data tables; however, if the manuscript intends a specific ICC or mixed-effect model for rater consistency, that exact inferential model should be locked before journal submission.

## Recommended next actions before submission

1. Populate `derived_data/workflow_total_time_cost_TEMPLATE.csv` with total human workflow time, total MAS workflow time, final item count, and nondefective item count; then regenerate Figure 4D-F.
2. Ask the first author or adjudication experts for a per-item defect adjudication table with columns such as `item_id`, `major_defect`, `critical_defect`, `defect_category`, `adjudicator_id`, and `final_decision`.
3. Decide whether Figure 1 will be submitted as the current separate panels or assembled into one integrated vector figure.
4. Lock the statistical model for Figure 2E rater agreement if journal reviewers are expected to request formal reliability statistics.
"""
    (DOCS/"update_full_review_after_source_data.md").write_text(review, encoding="utf-8")
    print(f"Wrote {len(manifest)} source-data workbooks to {SOURCE_DIR}")
    print(f"Wrote review report to {DOCS/'update_full_review_after_source_data.md'}")

if __name__ == "__main__":
    main()
