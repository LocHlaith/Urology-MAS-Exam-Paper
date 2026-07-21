"""抽样、调用模型并写入 MajorDefect critical defect 标注。

脚本用途：按题型比例抽取题目，调用 DeepSeek 标注 critical defects，并写入 Excel。
流程阶段：题库机器评价。
主要输入：题库 JSON、`prompts/evaluation/prompt_for_major_defects.txt`、`prompts/evaluation/major_defects.md`。
主要输出：抽样 JSON、模型标注 JSON、写入 D:J 七类缺陷列的 Excel。
重要边界：默认不会调用 DeepSeek API；只有 `annotate --run-api` 才会发起网络请求。
"""

from __future__ import annotations

import argparse
import json
import os
import random
import re
import sys
import time
import traceback
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

import requests

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from deepseek_env import load_dotenv
from project_paths import ENV_PATH, EVALUATION_PROMPT_DIR, OUTPUT_DIR, PROJECT_ROOT


# ===== 路径与常量 =====

CRITICAL_DEFECT_DIR = OUTPUT_DIR / "critical_defects"
SAMPLE_DIR = CRITICAL_DEFECT_DIR / "samples"
ANNOTATION_DIR = CRITICAL_DEFECT_DIR / "annotations"
LOG_DIR = CRITICAL_DEFECT_DIR / "logs"

DEFAULT_PROMPT_PATH = EVALUATION_PROMPT_DIR / "prompt_for_major_defects.txt"
DEFAULT_RUBRIC_PATH = EVALUATION_PROMPT_DIR / "major_defects.md"
DEFAULT_XLSX_PATH = CRITICAL_DEFECT_DIR / "critical_defects_AI标记.xlsx"

DEFECT_NAMES = {
    1: "题干缺陷",
    2: "选项缺陷",
    3: "格式缺陷",
    4: "评分缺陷",
    5: "公平性缺陷",
    6: "时效性缺陷",
    7: "结构缺陷",
}

STRIP_KEYS_FOR_MODEL = {
    "prototype",
    "fuzzywuzzy_doubt",
    "fuzzywuzzy_ratio_max",
    "sentencebert_doubt",
    "sentencebert_cosine_max",
    "3gram_doubt",
    "3gram_jaccard_max",
    "textstat_flesch_reading_ease",
    "QGEval",
    "ULM",
    "MajorDefects",
    "critical_defects",
}


# ===== 基础文件操作 =====

def ensure_dirs() -> None:
    for path in [CRITICAL_DEFECT_DIR, SAMPLE_DIR, ANNOTATION_DIR, LOG_DIR]:
        path.mkdir(parents=True, exist_ok=True)


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8-sig")


def read_json_list(path: Path) -> List[Dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if not isinstance(data, list):
        raise ValueError(f"JSON 顶层必须为 list: {path}")
    for i, item in enumerate(data):
        if not isinstance(item, dict):
            raise ValueError(f"{path} 第 {i} 项不是 dict")
    return data


def write_json(path: Path, obj: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")


# ===== 题目格式化 =====

def sorted_options(options: Any) -> List[str]:
    if not isinstance(options, dict):
        return []
    return [f"{key}. {options[key]}" for key in sorted(options.keys())]


def test_point_text(item: Dict[str, Any]) -> str:
    value = item.get("test_point")
    if value is None:
        return ""
    if isinstance(value, list):
        return ",".join(str(x) for x in value)
    return str(value)


def append_answer_block(lines: List[str], item: Dict[str, Any], idx: Optional[int] = None) -> None:
    suffix = "" if idx is None else str(idx)
    answer_key = f"answer{suffix}"
    analysis_key = f"analysis{suffix}"
    if answer_key in item:
        lines.append(f"参考答案：{item.get(answer_key)}")
    if analysis_key in item:
        lines.append(f"答案解析：{item.get(analysis_key)}")


def question_to_text(item: Dict[str, Any]) -> str:
    qtype = str(item.get("type", "")).upper()
    lines: List[str] = [f"题型：{qtype}", f"原始id：{item.get('id', '')}"]
    if item.get("test_point") is not None:
        lines.append(f"考点还原：{test_point_text(item)}")

    if qtype in {"A3", "A4"}:
        if item.get("case"):
            lines.append(f"病例摘要：{item.get('case')}")
        for i in range(1, 6):
            stem_key = f"stem{i}"
            option_key = f"options{i}"
            if stem_key not in item:
                continue
            lines.append(f"（{i}）{item.get(stem_key)}")
            lines.extend(sorted_options(item.get(option_key)))
            append_answer_block(lines, item, i)
        return "\n".join(lines)

    if qtype == "B":
        lines.extend(sorted_options(item.get("options")))
        for i in range(1, 6):
            stem_key = f"stem{i}"
            if stem_key not in item:
                continue
            lines.append(f"（{i}）{item.get(stem_key)}")
            append_answer_block(lines, item, i)
        return "\n".join(lines)

    if item.get("stem"):
        lines.append(str(item.get("stem")))
    lines.extend(sorted_options(item.get("options")))
    append_answer_block(lines, item, None)
    return "\n".join(lines)


def question_units(item: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Return independently scorable units.

    A3/A4 and B questions must never be submitted to the rater as one aggregate
    item: a defect in one subquestion must not automatically mark its siblings.
    Each returned unit contains the shared case/options only when they are needed
    to answer that particular subquestion.
    """
    qtype = str(item.get("type", "")).upper()
    common = [f"题型：{qtype}", f"原始id：{item.get('id', '')}"]
    if item.get("test_point") is not None:
        common.append(f"考点还原：{test_point_text(item)}")

    if qtype in {"A3", "A4"}:
        if item.get("case"):
            common.append(f"病例摘要：{item.get('case')}")
        units: List[Dict[str, Any]] = []
        for i in range(1, 6):
            stem_key = f"stem{i}"
            option_key = f"options{i}"
            if stem_key not in item:
                continue
            lines = [*common, f"小问：{i}", str(item.get(stem_key))]
            lines.extend(sorted_options(item.get(option_key)))
            append_answer_block(lines, item, i)
            units.append({"subquestion": i, "question_text": "\n".join(lines)})
        return units

    if qtype == "B":
        shared_options = sorted_options(item.get("options"))
        units = []
        for i in range(1, 6):
            stem_key = f"stem{i}"
            if stem_key not in item:
                continue
            lines = [*common, f"小问：{i}", *shared_options, str(item.get(stem_key))]
            append_answer_block(lines, item, i)
            units.append({"subquestion": i, "question_text": "\n".join(lines)})
        return units

    return [{"subquestion": None, "question_text": question_to_text(item)}]


def compact_for_model(
    item: Dict[str, Any],
    row: int,
    parent_row: int,
    subquestion: Optional[int],
    question_text: str,
) -> Dict[str, Any]:
    return {
        "row": row,
        "parent_row": parent_row,
        "subquestion": subquestion,
        "id": str(item.get("id", "")),
        "type": str(item.get("type", "")),
        "question_text": question_text,
    }


# ===== 按比例抽样 =====

def proportional_allocation(type_counts: Counter[str], sample_size: int) -> Dict[str, int]:
    total = sum(type_counts.values())
    if sample_size >= total:
        return dict(type_counts)

    raw = {key: type_counts[key] * sample_size / total for key in type_counts}
    alloc = {key: int(raw[key]) for key in type_counts}

    # 每个存在的题型尽量至少抽 1 题，避免小题型完全消失。
    for key in sorted(type_counts):
        if type_counts[key] > 0 and alloc[key] == 0 and sample_size >= len(type_counts):
            alloc[key] = 1

    while sum(alloc.values()) > sample_size:
        candidates = [k for k in alloc if alloc[k] > 0]
        key = min(candidates, key=lambda k: (raw[k] - int(raw[k]), alloc[k]))
        alloc[key] -= 1

    while sum(alloc.values()) < sample_size:
        candidates = [k for k in type_counts if alloc[k] < type_counts[k]]
        key = max(candidates, key=lambda k: (raw[k] - int(raw[k]), type_counts[k]))
        alloc[key] += 1

    return alloc


def sample_proportionally(items: List[Dict[str, Any]], sample_size: int, seed: int) -> List[Dict[str, Any]]:
    groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for item in items:
        groups[str(item.get("type", "")).upper()].append(item)

    counts = Counter({key: len(value) for key, value in groups.items()})
    alloc = proportional_allocation(counts, sample_size)
    rng = random.Random(seed)

    selected: List[Dict[str, Any]] = []
    for key in sorted(groups):
        pool = list(groups[key])
        rng.shuffle(pool)
        selected.extend(pool[: alloc.get(key, 0)])

    original_pos = {id(item): i for i, item in enumerate(items)}
    selected.sort(key=lambda item: original_pos[id(item)])
    return selected


# ===== DeepSeek API =====

class DeepSeekClient:
    def __init__(self, api_key: str, base_url: str, model: str, timeout: int = 180) -> None:
        self.api_key = api_key
        self.base_url = base_url.rstrip("/")
        self.model = model
        self.timeout = timeout

    def chat(self, system_prompt: str, user_prompt: str, temperature: float = 0.1) -> str:
        url = f"{self.base_url}/chat/completions"
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
        }
        payload = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            "temperature": temperature,
        }
        resp = requests.post(url, headers=headers, json=payload, timeout=self.timeout)
        if resp.status_code >= 400:
            raise RuntimeError(f"DeepSeek API HTTP {resp.status_code}: {resp.text[:2000]}")
        obj = resp.json()
        return obj["choices"][0]["message"]["content"]


_JSON_FENCE_RE = re.compile(r"```(?:json)?\s*([\s\S]*?)\s*```", re.IGNORECASE)


def extract_first_json_array(text: str) -> Any:
    text = (text or "").strip()
    fence = _JSON_FENCE_RE.search(text)
    if fence:
        return json.loads(fence.group(1).strip())
    start = text.find("[")
    if start == -1:
        raise ValueError("返回中未找到 JSON 数组")

    depth = 0
    in_str = False
    escape = False
    for i in range(start, len(text)):
        ch = text[i]
        if in_str:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_str = False
        else:
            if ch == '"':
                in_str = True
            elif ch == "[":
                depth += 1
            elif ch == "]":
                depth -= 1
                if depth == 0:
                    return json.loads(text[start:i + 1])
    raise ValueError("返回中未找到完整 JSON 数组")


def validate_annotation_rows(rows: Any) -> List[Dict[str, Any]]:
    if not isinstance(rows, list):
        raise ValueError("模型返回不是 list")
    out: List[Dict[str, Any]] = []
    for i, row in enumerate(rows):
        if not isinstance(row, dict):
            raise ValueError(f"第 {i} 行不是 dict")
        defects = row.get("defects")
        if not isinstance(defects, list) or len(defects) != 7:
            raise ValueError(f"第 {i} 行 defects 必须为 7 位 list")
        bits: List[int] = []
        for bit in defects:
            if bit not in [0, 1]:
                raise ValueError(f"第 {i} 行 defects 出现非 0/1: {defects}")
            bits.append(int(bit))
        reasons = row.get("reasons") or {}
        if not isinstance(reasons, dict):
            raise ValueError(f"第 {i} 行 reasons 不是 dict")
        out.append({
            "row": int(row["row"]),
            "id": str(row.get("id", "")),
            "defects": bits,
            "reasons": {str(k): str(v) for k, v in reasons.items()},
        })
    return out


def make_prompt(prompt_template: str, rubric: str, questions: List[Dict[str, Any]]) -> str:
    questions_json = json.dumps(questions, ensure_ascii=False, indent=2)
    return (
        prompt_template
        .replace("{rubric}", rubric.strip())
        .replace("{questions_json}", questions_json)
    )


# ===== Excel 写入 =====

def write_annotations_to_xlsx(
    xlsx_path: Path,
    sheet_name: str,
    annotations: List[Dict[str, Any]],
    start_row: int = 2,
) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.comments import Comment
    except ModuleNotFoundError as exc:
        raise RuntimeError("写入 xlsx 需要 openpyxl；请先安装 requirements.txt 中的依赖。") from exc

    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.cell(1, 1).value = "序号"
        ws.cell(1, 2).value = "题目"
        ws.cell(1, 3).value = "解析"
    else:
        ws = wb[sheet_name]

    for i in range(1, 8):
        ws.cell(1, i + 3).value = i
        ws.cell(1, i + 10).value = f"{i}理由"

    for ann in annotations:
        # ``parent_row`` is retained for model outputs expanded from one A3/A4/B
        # source item.  An expanded review workbook may instead supply an explicit
        # ``excel_row`` for each subquestion.  In either case, retain the subquestion
        # prefix in the reason so that independent judgments are not collapsed into
        # a generic statement.
        source_row = int(ann.get("excel_row", ann.get("parent_row", ann["row"])))
        row = start_row + source_row - 1
        defects = ann["defects"]
        reasons = ann.get("reasons", {})
        subquestion = ann.get("subquestion")
        for defect_idx, bit in enumerate(defects, start=1):
            cell = ws.cell(row=row, column=3 + defect_idx)
            reason_cell = ws.cell(row=row, column=10 + defect_idx)
            # Do not erase a positive flag written for an earlier sibling unit.
            cell.value = int(bool(cell.value) or bool(bit))
            if bit:
                reason = reasons.get(str(defect_idx), "")
                label = DEFECT_NAMES[defect_idx]
                prefix = "" if subquestion is None else f"第{subquestion}小问："
                reason_text = f"{defect_idx} {label}: {prefix}{reason}".strip()
                # More than one expanded unit may target one source row. Preserve
                # each individual rationale rather than overwriting a sibling's.
                old_reason = str(reason_cell.value or "").strip()
                reason_cell.value = "\n".join(x for x in [old_reason, reason_text] if x)
                cell.comment = Comment(reason_text, "DeepSeek")
            else:
                reason_cell.value = None
                cell.comment = None

    wb.save(xlsx_path)


# ===== TXT 导入 =====

MAJOR_RE = re.compile(r"^Major Defects：(.*)；标记：\[([01](?:,[01]){6})\]$", re.M)
ITEM_RE = re.compile(r"(\d+)\s+([^（；]+)（理由：([^）]+)）")


def parse_major_defects_txt(path: Path) -> List[Dict[str, Any]]:
    text = path.read_text(encoding="utf-8-sig")
    rows: List[Dict[str, Any]] = []
    for row_number, match in enumerate(MAJOR_RE.finditer(text), start=1):
        body = match.group(1).strip()
        defects = [int(x) for x in match.group(2).split(",")]
        reasons: Dict[str, str] = {}
        if body != "无":
            for item in ITEM_RE.finditer(body):
                reasons[item.group(1)] = item.group(3).strip()
        rows.append({"row": row_number, "id": "", "defects": defects, "reasons": reasons})
    if len(rows) != 50:
        raise ValueError(f"{path} 解析出 {len(rows)} 行 Major Defects，预期 50")
    return rows


# ===== 子命令 =====

def cmd_sample(args: argparse.Namespace) -> None:
    ensure_dirs()
    items = read_json_list(Path(args.input_json))
    selected = sample_proportionally(items, args.sample_size, args.seed)
    write_json(Path(args.output_json), selected)
    print(f"[OK] sampled {len(selected)} / {len(items)} -> {args.output_json}")


def cmd_annotate(args: argparse.Namespace) -> None:
    ensure_dirs()
    sample_items = read_json_list(Path(args.sample_json))
    prompt_template = read_text(Path(args.prompt))
    rubric = read_text(Path(args.rubric))
    model_items: List[Dict[str, Any]] = []
    for parent_row, item in enumerate(sample_items, start=1):
        stripped = {k: v for k, v in item.items() if k not in STRIP_KEYS_FOR_MODEL}
        for unit in question_units(stripped):
            model_items.append(compact_for_model(
                item,
                row=len(model_items) + 1,
                parent_row=parent_row,
                subquestion=unit["subquestion"],
                question_text=unit["question_text"],
            ))
    batches = [model_items[i:i + args.batch_size] for i in range(0, len(model_items), args.batch_size)]

    if not args.run_api:
        preview_path = Path(args.output_json).with_suffix(".prompt_preview.txt")
        prompt = make_prompt(prompt_template, rubric, batches[0] if batches else [])
        preview_path.parent.mkdir(parents=True, exist_ok=True)
        preview_path.write_text(prompt, encoding="utf-8")
        print(f"[DRY-RUN] 未调用 API。首批 prompt 已写入 {preview_path}")
        print("[DRY-RUN] 若确认要调用 DeepSeek，请显式增加 annotate --run-api。")
        return

    load_dotenv(ENV_PATH)
    api_key = (os.getenv(args.api_key_env) or "").strip()
    if not api_key:
        raise RuntimeError(f"未读取到 {args.api_key_env}")
    model = (os.getenv("DEEPSEEK_MODEL") or "deepseek-reasoner").strip()
    base_url = (os.getenv("DEEPSEEK_BASE_URL") or "https://api.deepseek.com").strip()
    client = DeepSeekClient(api_key=api_key, base_url=base_url, model=model, timeout=args.timeout)

    all_rows: List[Dict[str, Any]] = []
    system_prompt = "你是严谨的医学考试题库审题人，只输出符合要求的 JSON。"
    metadata_by_row = {entry["row"]: entry for entry in model_items}
    for batch_index, batch in enumerate(batches, start=1):
        prompt = make_prompt(prompt_template, rubric, batch)
        log_path = LOG_DIR / f"major_defects_batch_{batch_index:04d}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        try:
            response = client.chat(system_prompt=system_prompt, user_prompt=prompt, temperature=0.1)
            log_path.write_text(response, encoding="utf-8")
            rows = validate_annotation_rows(extract_first_json_array(response))
            for ann in rows:
                metadata = metadata_by_row.get(ann["row"])
                if metadata is None:
                    raise ValueError(f"模型返回了不存在的 row: {ann['row']}")
                ann["parent_row"] = metadata["parent_row"]
                ann["subquestion"] = metadata["subquestion"]
            all_rows.extend(rows)
            print(f"[OK] batch {batch_index}/{len(batches)} rows={len(rows)}")
            time.sleep(args.sleep)
        except Exception:
            log_path.write_text(traceback.format_exc(), encoding="utf-8")
            raise

    write_json(Path(args.output_json), all_rows)
    print(f"[OK] annotations -> {args.output_json}")


def cmd_write_xlsx(args: argparse.Namespace) -> None:
    ensure_dirs()
    annotations = validate_annotation_rows(json.loads(Path(args.annotations_json).read_text(encoding="utf-8")))
    write_annotations_to_xlsx(Path(args.xlsx), args.sheet, annotations, start_row=args.start_row)
    print(f"[OK] wrote {len(annotations)} rows -> {args.xlsx} [{args.sheet}]")


def cmd_import_txt(args: argparse.Namespace) -> None:
    ensure_dirs()
    annotations = parse_major_defects_txt(Path(args.txt))
    write_annotations_to_xlsx(Path(args.xlsx), args.sheet, annotations, start_row=args.start_row)
    print(f"[OK] imported {len(annotations)} rows from {args.txt} -> {args.xlsx} [{args.sheet}]")


# ===== 命令行入口 =====

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="MajorDefect critical defects 标注链路。")
    sub = parser.add_subparsers(dest="command", required=True)

    p_sample = sub.add_parser("sample", help="按题型比例抽取题目。")
    p_sample.add_argument("--input-json", required=True)
    p_sample.add_argument("--output-json", default=str(SAMPLE_DIR / "major_defect_sample.json"))
    p_sample.add_argument("--sample-size", type=int, default=50)
    p_sample.add_argument("--seed", type=int, default=20260711)
    p_sample.set_defaults(func=cmd_sample)

    p_annotate = sub.add_parser("annotate", help="调用 DeepSeek 生成 MajorDefect 标注；默认 dry-run。")
    p_annotate.add_argument("--sample-json", required=True)
    p_annotate.add_argument("--output-json", default=str(ANNOTATION_DIR / "major_defect_annotations.json"))
    p_annotate.add_argument("--prompt", default=str(DEFAULT_PROMPT_PATH))
    p_annotate.add_argument("--rubric", default=str(DEFAULT_RUBRIC_PATH))
    p_annotate.add_argument("--batch-size", type=int, default=20)
    p_annotate.add_argument("--timeout", type=int, default=240)
    p_annotate.add_argument("--sleep", type=float, default=0.6)
    p_annotate.add_argument("--api-key-env", default="DEEPSEEK_API_KEY_003")
    p_annotate.add_argument("--run-api", action="store_true", help="显式允许调用 DeepSeek API。")
    p_annotate.set_defaults(func=cmd_annotate)

    p_write = sub.add_parser("write-xlsx", help="将模型标注 JSON 写入 Excel 的 1-7 缺陷列。")
    p_write.add_argument("--annotations-json", required=True)
    p_write.add_argument("--xlsx", default=str(DEFAULT_XLSX_PATH))
    p_write.add_argument("--sheet", required=True)
    p_write.add_argument("--start-row", type=int, default=2)
    p_write.set_defaults(func=cmd_write_xlsx)

    p_import = sub.add_parser("import-txt", help="将人工 TXT 中的 Major Defects 行写入 Excel。")
    p_import.add_argument("--txt", required=True)
    p_import.add_argument("--xlsx", default=str(DEFAULT_XLSX_PATH))
    p_import.add_argument("--sheet", required=True)
    p_import.add_argument("--start-row", type=int, default=2)
    p_import.set_defaults(func=cmd_import_txt)

    return parser.parse_args()


def main() -> None:
    args = parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
